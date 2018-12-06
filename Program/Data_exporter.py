# Imports
# --> Modules
import pandas as pd
import numpy as np
import openpyxl
import time
import pulp

# --> Module Classes
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# --> Home made files
import Chart_creator as ChC
import Converters as CONV



def save_data(input_data, Bay_Assignment, solve_status):
    print ('Exporting data to Excel: ...')
    start_time_export = time.time()

    # Converting input_data from list to dataframe (for later)
    input_dataframe  = CONV.inputs_list2dataframe(input_data)
    output_dataframe = pd.DataFrame(input_dataframe)

    # Exporting inputs to a csv file (for eventual later use)
    input_dataframe.to_csv('./outputs/Generated Inputs.csv')

    if solve_status.lower() == 'optimal':
        # Adding assigned bay to flights
        bay_assignment = list(np.zeros(len(input_data)))
        
        for decision_variable in Bay_Assignment.iter_binary_vars():
            if int(decision_variable.solution_value) and decision_variable.name.find('x') != -1:
                
                dv, var_type, flight_index, bay = decision_variable.name.split('_')
                bay_assignment[int(flight_index)] = bay

            if (decision_variable.name.find('v') != -1 or  decision_variable.name.find('w')!= -1 or  decision_variable.name.find('u')!= -1) and int(decision_variable.solution_value):
                print(decision_variable.name, decision_variable.solution_value)
                #pass


        output_dataframe['Bay'] = bay_assignment

        # Checking for towings:
        long_stays = output_dataframe[output_dataframe['long stay']]

        arrivals   = long_stays[long_stays['move type'] == 'Arr' ].reset_index()
        parkings   = long_stays[long_stays['move type'] == 'Park'].reset_index()
        departures = long_stays[long_stays['move type'] == 'Dep' ].reset_index()



        towings_ = []  
        for i in range(len(list(arrivals['Bay']))):
            towing_arr_park = 'NO'
            towing_park_dep = 'NO'
            
            if arrivals['Bay'].iloc[i] != parkings['Bay'].iloc[i]:
                towing_arr_park = 'YES'
            
            if parkings['Bay'].iloc[i] != departures['Bay'].iloc[i]:
                towing_park_dep = 'YES'

            towings_.append({'Fl No. Arrival' : arrivals['Fl No. Arrival'].iloc[i],
                             'Arrival Bay'    : arrivals  ['Bay'].iloc[i]         ,
                             'Park Bay'       : parkings  ['Bay'].iloc[i]         ,
                             'Departure Bay'  : departures['Bay'].iloc[i]         ,
                             'Arr -> Park'    : towing_arr_park                   , 
                             'Park -> Dep'    : towing_park_dep
                            })


        #print (towings)
        
        towings_dataframe = pd.DataFrame.from_records(towings_)
        if len(towings_dataframe) > 0:
            towings_dataframe = towings_dataframe[['Fl No. Arrival'  ,
                                                   'Arrival Bay'     ,
                                                   'Park Bay'        ,
                                                   'Departure Bay'   ,
                                                   'Arr -> Park'     ,
                                                   'Park -> Dep'     ]]
        else:
            towings_dataframe[['Fl No. Arrival'  ,
                               'Arrival Bay'     ,
                               'Park Bay'        ,
                               'Departure Bay'   ,
                               'Arr -> Park'     ,
                               'Park -> Dep'     ]] = [0,0,0,0,0,0]

        # Exporting all data to excel
        export_2_excel(input_dataframe, output_dataframe, towings_dataframe)

        print ('Exporting data to Excel: DONE (' + str(round(time.time()-start_time_export, 3)) +' seconds)\n')



        # Creating Charts
        print ('Generating charts: ...')
        start_time_charts = time.time()

        ChC.generate_charts(input_dataframe, output_dataframe, towings_dataframe, solve_status)
        
        print ('Generating charts: DONE (' + str(round(time.time()-start_time_charts, 3)) +' seconds)\n')
    
    
    return output_dataframe

def export_2_excel(input_dataframe, output_dataframe, towings_dataframe):

        # Creating the excel file
        wb = Workbook()
        

        # Creating needed sheets
        wb.create_sheet(index = 2, title = 'Inputs')
        wb.create_sheet(index = 3, title = 'Outputs')
        wb.create_sheet(index = 4, title = 'Towings')
        wb.remove_sheet(wb['Sheet'])
        

        # Adding inputs to first sheet
        for row in dataframe_to_rows(input_dataframe, index=False):
            wb['Inputs'].append(row)


        # Adding outputs to second sheet
        for row in dataframe_to_rows(output_dataframe, index=False):
            wb['Outputs'].append(row)       

        
        # Adding towings to third sheet
        for row in dataframe_to_rows(towings_dataframe, index=False):
            wb['Towings'].append(row)         

        # Saving file
        save_as = './outputs/Bay Assignment.xlsx'
        wb.save(filename = save_as)
        print ('  ---> Saved as "' + save_as + '"')
















        
